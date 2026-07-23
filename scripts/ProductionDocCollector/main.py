from __future__ import annotations

import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook


# ============================================================
# CONFIGURAZIONE
# ============================================================

VALID_EXTENSIONS = {".pdf", ".stp", ".step", ".stl"}

OUTPUT_FOLDER_NAME = "ProductionDocCollector_Output"

# Colonne Excel attese. La ricerca e' tollerante a maiuscole/minuscole,
# spazi e punto finale, quindi accetta ad esempio: Cod, Cod., COD.
COD_HEADER_ALIASES = {
    "cod",
    "codice",
}

OPTIONAL_HEADER_ALIASES = {
    "optionalcolumn2",
    "optional2",
}

def build_production_doc_path() -> Path:
    """
    Cerca automaticamente la cartella 'Production Doc'
    all'interno della cartella IIT dell'utente corrente.

    Compatibile con:
    - iCub Tech - Documents
    - MESH - Silo Mechanics
    - eventuali futuri cambi di struttura
    """

    current_user = Path.home().name

    root = (
        Path("C:/Users")
        / current_user
        / "Fondazione Istituto Italiano Tecnologia"
    )

    if not root.exists():
        raise FileNotFoundError(
            f"IIT folder not found:\n{root}"
        )

    matches = []

    for path in root.rglob("Production Doc"):
        if path.is_dir():
            matches.append(path)

    if len(matches) == 0:
        raise FileNotFoundError(
            "No 'Production Doc' folder found."
        )

    if len(matches) > 1:
        message = (
            "More than one 'Production Doc' folder found:\n\n"
            + "\n".join(str(p) for p in matches)
        )
        raise RuntimeError(message)

    return matches[0]


# ============================================================
# UTILITY GENERALI
# ============================================================

def normalize_header(value) -> str:
    """
    Normalizza gli header Excel:
    - minuscolo
    - rimuove spazi, punti, underscore e trattini

    Esempi:
    "Cod." -> "cod"
    "Optional column 2" -> "optionalcolumn2"
    "OPTIONAL_COLUMN_2" -> "optionalcolumn2"
    """
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = re.sub(r"[\s._\-]+", "", text)
    return text


def cell_to_string(value) -> str:
    """
    Converte una cella Excel in stringa pulita.
    Evita il classico problema di Excel/Python con valori numerici tipo 123.0.
    """
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()

    return str(value).strip()

def extract_revision(code: str) -> Optional[str]:
    """
    Estrae la revisione finale dal codice principale.

    Esempi validi:
    IC_021_P_328_rev_0  -> 0
    IC_021_P_328_rev_A1 -> A1
    IC_021_P_328_REV_B9 -> B9

    Le revisioni vengono trattate come stringhe, non come numeri.
    Quindi non viene fatta alcuna logica di incremento/confronto.
    """
    match = re.search(r"_rev_([A-Z0-9]+)$", code, flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1)


def remove_existing_revision(code: str) -> str:
    """
    Se un optional arrivasse gia' con _rev_X, lo rimuove.
    Poi verra' aggiunta la revisione del codice principale.
    """
    return re.sub(r"_rev_[A-Z0-9]+$", "", code, flags=re.IGNORECASE)


# ============================================================
# LETTURA EXCEL
# ============================================================

def find_column_indexes(ws) -> Tuple[int, Optional[int]]:
    """
    Trova le colonne:
    - Cod. / Cod / codice
    - Optional column 2 / optional2 / optional

    Ritorna indici 1-based, come richiesto da openpyxl.
    La colonna Cod e' obbligatoria.
    La colonna Optional column 2 e' opzionale.
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1))

    cod_col = None
    optional_col = None

    for cell in header_row:
        header = normalize_header(cell.value)

        if header in COD_HEADER_ALIASES:
            cod_col = cell.column

        if header in OPTIONAL_HEADER_ALIASES:
            optional_col = cell.column

    if cod_col is None:
        raise ValueError(
            "Mandatory column not found. "
            "I was expecting a column called 'Cod.' or 'Cod'."
        )

    return cod_col, optional_col


def build_codes_from_excel(excel_path: Path) -> Tuple[List[str], List[str]]:
    """
    Costruisce la lista dei codici da cercare.

    Per ogni riga Excel:
    1. aggiunge il codice principale dalla colonna Cod.
    2. se Optional column 2 e' valorizzata:
       - estrae la revisione dal codice principale
       - rimuove una eventuale revisione gia' presente dall'optional
       - costruisce optional_rev_<revisione del principale>
       - aggiunge anche quello alla lista

    P e G sono trattati come codici separati.
    Se manca il P, finisce in missing.txt.
    Se manca il G, finisce in missing.txt.
    Se mancano entrambi, finiscono entrambi in missing.txt.
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    cod_col, optional_col = find_column_indexes(ws)

    codes: List[str] = []
    warnings: List[str] = []

    for row_idx in range(2, ws.max_row + 1):
        main_code = cell_to_string(ws.cell(row=row_idx, column=cod_col).value)

        if not main_code:
            continue

        codes.append(main_code)

        optional_code = ""
        if optional_col is not None:
            optional_code = cell_to_string(ws.cell(row=row_idx, column=optional_col).value)

        if optional_code:
            revision = extract_revision(main_code)

            if revision is None:
                warnings.append(
                    f"Riga {row_idx}: impossibile estrarre la revisione da '{main_code}'. "
                    f"Optional ignorato: '{optional_code}'."
                )
                continue

            optional_base = remove_existing_revision(optional_code)
            optional_with_revision = f"{optional_base}_rev_{revision}"
            codes.append(optional_with_revision)

    # Rimuove duplicati mantenendo l'ordine originale.
    unique_codes: List[str] = []
    seen = set()

    for code in codes:
        key = code.lower()
        if key not in seen:
            unique_codes.append(code)
            seen.add(key)

    return unique_codes, warnings


# ============================================================
# INDICIZZAZIONE FILE
# ============================================================

def index_production_doc_files(production_doc_path: Path, output_path: Path) -> Dict[str, List[Path]]:
    """
    Scansiona Production Doc ricorsivamente una sola volta.

    Cerca in:
    - Production Doc
    - tutte le sottocartelle
    - sottocartelle delle sottocartelle, ecc.

    Crea un indice case-insensitive basato sul nome file senza estensione.

    Esempio:
    IC_021_P_328_rev_A3.pdf  -> chiave: ic_021_p_328_rev_a3
    IC_021_P_328_rev_A3.step -> chiave: ic_021_p_328_rev_a3

    Nota: il matching e' esatto sul nome senza estensione.
    Quindi IC_021_P_328_rev_A3_extra.pdf NON viene associato a IC_021_P_328_rev_A3.
    """
    index: Dict[str, List[Path]] = defaultdict(list)

    for file_path in production_doc_path.rglob("*"):
        if not file_path.is_file():
            continue

        # Non indicizzare mai la cartella Output.
        try:
            file_path.relative_to(output_path)
            continue
        except ValueError:
            pass

        if file_path.suffix.lower() not in VALID_EXTENSIONS:
            continue

        key = file_path.stem.lower()
        index[key].append(file_path)

    return index


# ============================================================
# OUTPUT
# ============================================================

def reset_output_folder(output_path: Path) -> None:
    """
    Svuota la cartella Output senza cancellarla.

    Questa versione è molto più robusta con
    SharePoint / OneDrive sincronizzati.
    """

    output_path.mkdir(parents=True, exist_ok=True)

    for item in output_path.iterdir():
        try:
            if item.is_file():
                item.unlink()

            elif item.is_dir():
                shutil.rmtree(item)

        except Exception as exc:
            print(
                f"WARNING: impossible to eliminate "
                f"'{item}': {exc}"
            )


def unique_destination_path(output_path: Path, file_name: str) -> Path:
    """
    Evita collisioni nel caso raro in cui due sottocartelle contengano file con lo stesso nome.
    In quel caso aggiunge __1, __2, ecc.
    """
    destination = output_path / file_name

    if not destination.exists():
        return destination

    stem = destination.stem
    suffix = destination.suffix
    counter = 1

    while True:
        candidate = output_path / f"{stem}__{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def copy_found_files(
    codes: List[str],
    file_index: Dict[str, List[Path]],
    output_path: Path,
) -> Tuple[int, List[str], List[str]]:
    """
    Per ogni codice:
    - se trova uno o piu' file, li copia tutti in Output
    - se non trova nessun file, aggiunge il codice a missing_codes

    Ritorna:
    - numero file copiati
    - lista codici mancanti
    - warnings/errori di copia
    """
    found_files_count = 0
    missing_codes: List[str] = []
    warnings: List[str] = []

    for code in codes:
        key = code.lower()
        found_files = file_index.get(key, [])

        if not found_files:
            missing_codes.append(code)
            continue

        for source_file in found_files:
            destination = unique_destination_path(output_path, source_file.name)
            try:
                shutil.copy2(source_file, destination)
                found_files_count += 1
            except Exception as exc:
                warnings.append(
                    f"Error copying '{source_file}' -> '{destination}': {exc}"
                )

    return found_files_count, missing_codes, warnings


def write_missing_txt(
    output_path: Path,
    production_doc_path: Path,
    excel_path: Path,
    total_codes: int,
    found_files_count: int,
    missing_codes: List[str],
    warnings: List[str],
) -> None:
    """
    Scrive Output/missing.txt con riepilogo e lista mancanti.
    """
    missing_file = output_path / "missing.txt"

    with missing_file.open("w", encoding="utf-8") as f:
        f.write("Production Doc path:\n")
        f.write(f"{production_doc_path}\n\n")

        f.write("Excel file:\n")
        f.write(f"{excel_path}\n\n")

        f.write(f"Total codes checked: {total_codes}\n")
        f.write(f"Found files: {found_files_count}\n")
        f.write(f"Missing codes: {len(missing_codes)}\n")
        f.write("\n")
        f.write("--------------------------------\n")
        f.write("Missing codes list\n")
        f.write("--------------------------------\n\n")

        if missing_codes:
            for code in missing_codes:
                f.write(f"{code}\n")
        else:
            f.write("No missing codes.\n")

        if warnings:
            f.write("\n")
            f.write("--------------------------------\n")
            f.write("Warnings\n")
            f.write("--------------------------------\n\n")

            for warning in warnings:
                f.write(f"{warning}\n")


# ============================================================
# PROCESSO PRINCIPALE
# ============================================================

def run_process(excel_path: Path) -> Tuple[int, int, Path]:
    """
    Esegue tutto il workflow:
    1. costruisce e verifica il path Production Doc
    2. resetta Output
    3. legge Excel e costruisce codici
    4. indicizza Production Doc ricorsivamente
    5. copia file trovati
    6. scrive missing.txt
    """
    production_doc_path = build_production_doc_path()

    if not production_doc_path.exists():
        raise FileNotFoundError(
            "Path not matching.\n\n"
            f"Expected path:\n{production_doc_path}"
        )

    output_path = (
        Path.home()
        / "Documents"
        / OUTPUT_FOLDER_NAME
    )
    
    #output_path = production_doc_path / OUTPUT_FOLDER_NAME

    codes, excel_warnings = build_codes_from_excel(excel_path)

    reset_output_folder(output_path)

    file_index = index_production_doc_files(
        production_doc_path=production_doc_path,
        output_path=output_path,
    )

    found_files_count, missing_codes, copy_warnings = copy_found_files(
        codes=codes,
        file_index=file_index,
        output_path=output_path,
    )

    all_warnings = excel_warnings + copy_warnings

    write_missing_txt(
        output_path=output_path,
        production_doc_path=production_doc_path,
        excel_path=excel_path,
        total_codes=len(codes),
        found_files_count=found_files_count,
        missing_codes=missing_codes,
        warnings=all_warnings,
    )

    return found_files_count, len(missing_codes), output_path

def main():

    root = tk.Tk()
    root.withdraw()  # nasconde la finestra principale

    excel_file = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm"),
            ("All files", "*.*"),
        ],
    )

    if not excel_file:
        return

    try:
        found_count, missing_count, output_path = run_process(
            excel_path=Path(excel_file)
        )

        messagebox.showinfo(
            "Completato",
            f"Found files: {found_count}\n"
            f"Missing codes: {missing_count}\n\n"
            f"Output:\n{output_path}"
        )

    except Exception as exc:
        import traceback
        traceback.print_exc()

        messagebox.showerror(
            "Errore",
            f"{exc}"
        )
        
if __name__ == "__main__":
    main()
